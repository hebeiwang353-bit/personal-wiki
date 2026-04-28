"""
从各种格式文件中提取纯文本。
"""

from pathlib import Path

MAX_CHARS = 8000  # 单文件送给 LLM 的最大字符数


def extract(path: Path) -> str | None:
    """返回提取的文本，失败或空白则返回 None"""
    ext = path.suffix.lower()
    try:
        text = _dispatch(path, ext)
        if text:
            text = text.strip()
            return text[:MAX_CHARS] if len(text) > MAX_CHARS else text
        return None
    except Exception:
        return None


def _dispatch(path: Path, ext: str) -> str | None:
    if ext == ".pdf":
        return _pdf(path)
    if ext in (".doc", ".docx"):
        return _docx(path)
    if ext in (".xls", ".xlsx"):
        return _excel(path)
    if ext == ".csv":
        return path.read_text(errors="ignore")[:MAX_CHARS]
    if ext in (".html", ".htm"):
        return _html(path)
    # 纯文本类
    if ext in (".txt", ".md", ".markdown", ".py", ".js", ".ts", ".java",
               ".go", ".rs", ".swift", ".json", ".yaml", ".yml", ".rtf"):
        return path.read_text(errors="ignore")
    return None


def _pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    pages = []
    for page in reader.pages[:30]:  # 最多前 30 页
        t = page.extract_text()
        if t:
            pages.append(t)
    return "\n".join(pages)


def _docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _excel(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets[:3]:  # 最多前 3 个 sheet
        for row in ws.iter_rows(max_row=200, values_only=True):
            row_str = "\t".join(str(c) if c is not None else "" for c in row)
            if row_str.strip():
                lines.append(row_str)
    return "\n".join(lines)


def _html(path: Path) -> str:
    import re
    raw = path.read_text(errors="ignore")
    # 去掉标签，保留文字
    text = re.sub(r"<[^>]+>", " ", raw)
    text = re.sub(r"\s+", " ", text)
    return text
